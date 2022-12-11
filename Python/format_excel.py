import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from datetime import datetime

dia = str(datetime.now())[:10]

df = pd.DataFrame({
    'A': [1, 2, 3, 4, 5, 6],
    'B': [7, 8, 9, 10, 11, 12],
    'C': [13, 14, 15, 16, 17, 18],
    'D': [19, 20, 21, 22, 23, 24]
})

initial_row = 2

# Dataframe Base
df.to_excel('./Datos/01.df_base.xlsx')

# Dataframe sin indices
df.to_excel('./Datos/02.df_without_index.xlsx', index=False)

# Definir el nombre de la hoja
with pd.ExcelWriter('./Datos/03.df_sheetname.xlsx') as writer:
    df.to_excel(writer, sheet_name='Sheet 1', index=False, startrow=initial_row)

# Editar estilos
wb = load_workbook("./Datos/03.df_sheetname.xlsx")
ws = wb['Sheet 1']

cell = ws['A1']
cell.font = Font(bold=True, size=16)
cell.alignment = Alignment(horizontal='left')

cell.value = f'Hola, mundo {dia}.'

for row in ws.iter_rows(min_row=(initial_row + 1), max_col=df.shape[1], max_row=(initial_row + 1)):
    for cell in row:
        cell.fill = PatternFill("solid", fgColor="053c5e")
        cell.font = Font(color='FFFFFF', bold=True)

for row in ws.iter_rows(min_row=(initial_row + 2), max_col=df.shape[1], max_row=(initial_row + df.shape[0] + 1)):
    for cell in row:
        # Establece el formato de cada celda
        cell.alignment = Alignment(horizontal="center")

        cell.border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )

        if (cell.value > 10) & (cell.value % 2 == 0):
            cell.fill = PatternFill("solid", fgColor="FF0000")
            cell.font = Font(color='FFFFFF', bold=True)

wb.save("./Datos/04.df_styles.xlsx")